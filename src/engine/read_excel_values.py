import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('/Users/ianlioubachevskii/Claud/Hostplus_Retirement_Portfolios_Backtest.xlsx', data_only=True)
    ws = wb['ENGINE']
except Exception as e:
    print(f"Failed to load workbook: {e}")
    sys.exit(1)

rebalances = []
for row_idx in range(4, 300): # assuming max 300 rows
    date = ws.cell(row=row_idx, column=18).value
    if date is None:
        continue
    is_rebal = ws.cell(row=row_idx, column=31).value
    if is_rebal == 'Y':
        if hasattr(date, 'strftime'):
            rebalances.append(date.strftime('%Y-%m'))
        else:
            rebalances.append(str(date))

print(f"Total Rebalances in User's Excel: {len(rebalances)}")
print(rebalances)

# Let's also print some rows to verify values
print("\nFirst 5 rows of Balanced values:")
for row_idx in range(4, 9):
    date = ws.cell(row=row_idx, column=18).value
    if date is None:
        break
    if hasattr(date, 'strftime'): date = date.strftime('%Y-%m')
    
    growth_open = ws.cell(row=row_idx, column=20).value
    cpi_open = ws.cell(row=row_idx, column=21).value
    payment = ws.cell(row=row_idx, column=27).value
    growth_ret = ws.cell(row=row_idx, column=23).value
    cpi_ret = ws.cell(row=row_idx, column=24).value
    weight = ws.cell(row=row_idx, column=30).value
    print(f"{date}: Open=({growth_open}, {cpi_open}) Ret=({growth_ret}, {cpi_ret}) Pay={payment} Wt={weight}")
